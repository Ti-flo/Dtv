import os
import pandas as pd
from statistics import mean
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from datetime import datetime

# === 📁 CONFIGURATION DES DOSSIERS ===
# Répertoire du script actuel (DataCrossing)
script_folder = os.path.dirname(os.path.realpath(__file__))

# Répertoire des relevés
releves_folder = os.path.abspath(os.path.join(script_folder, "..", "Releves", "Suivis_par_categorie", "Ressources"))

# Fichier de synthèse de base
synth_path = os.path.abspath(os.path.join(script_folder, "..", "Releves", "Templates", "Synthese_Prix_Ressources.xlsx"))

# === 📂 FILTRAGE DES FICHIERS DE RELEVÉ ===
# On ne garde que les fichiers Excel qui commencent par "Suivi_Prix_Ressources_2" et qui finissent par ".xlsx"
files = [f for f in os.listdir(releves_folder) if f.startswith("Suivi_Prix_Ressources_2") and f.endswith(".xlsx")]
files.sort()

price_data = {}         # Contiendra toutes les valeurs par nom de ressource
last_file_data = {}     # Contiendra uniquement les prix du dernier fichier (pour "Prix Bas Actuel")

moyens_globaux = {}     # Liste de tous les "Prix Moyen" par ressource
mins_globaux = {}       # Tous les "Prix Min"
maxs_globaux = {}       # Tous les "Prix Max"

# === 📊 LECTURE DES FICHIERS DE RELEVÉ ===
for file in files:
    file_path = os.path.join(releves_folder, file)
    df = pd.read_excel(file_path)

    # Sécurité : on vérifie que la colonne "Prix Moyen" est bien présente
    if "Prix Moyen" not in df.columns:
        print(f"⚠️  Fichier ignoré : '{file}' n'a pas de colonne 'Prix Moyen'")
        continue

    for _, row in df.iterrows():
        nom = row["Nom"].strip()
        moyen = row["Prix Moyen"]
        mini = row["Prix Min"]
        maxi = row["Prix Max"]

        if nom not in price_data:
            price_data[nom] = {"last_min": None, "last_moyen": None, "last_max": None}
        if nom not in moyens_globaux:
            moyens_globaux[nom] = []
            mins_globaux[nom] = []
            maxs_globaux[nom] = []
            
        # Ajout aux historiques globaux
        moyens_globaux[nom].append(moyen)
        mins_globaux[nom].append(mini)
        maxs_globaux[nom].append(maxi)

        # Si on est sur le dernier fichier, on sauvegarde aussi le "Prix x1"
        if file == files[-1]:
            price_data[nom]["last_min"] = mini
            price_data[nom]["last_moyen"] = moyen
            price_data[nom]["last_max"] = maxi

# === 🧮 MISE À JOUR DU FICHIER DE SYNTHÈSE ===
wb_synth = load_workbook(synth_path)
ws_synth = wb_synth.active

# Crée une bordure noire à droite uniquement
right_border = Border(right=Side(style='thin', color='000000'))

# Colonnes où on applique une bordure à droite (A, C, F, G, I, K, Q, R, V, W)
colonnes_bordure_droite = [1, 3, 6, 7, 9, 11, 17, 18, 22, 23]

for row in ws_synth.iter_rows(min_row=2):
    nom = row[0].value.strip() if row[0].value else None

    if nom in price_data:
        row[3].value = price_data[nom]["last_min"]                    # Bas Actuel
        row[4].value = price_data[nom]["last_moyen"]                 # Moyen Actuel
        row[5].value = price_data[nom]["last_max"]                   # Haut Actuel

    if nom in moyens_globaux:
        row[6].value = round(mean(moyens_globaux[nom]), 2)           # Moyen dans le temps
        row[7].value = min(moyens_globaux[nom])                      # Moyen le plus bas
        row[8].value = max(moyens_globaux[nom])                      # Moyen le plus haut

    if nom in mins_globaux:
        row[9].value = min(mins_globaux[nom])                        # Plus Bas

    if nom in maxs_globaux:
        row[10].value = max(maxs_globaux[nom])                       # Plus Haut

    # 🔲 Appliquer une bordure droite noire aux colonnes spécifiées
    for col_index in colonnes_bordure_droite:
        if col_index <= len(row):  # sécurité
            cell = row[col_index - 1]  # car les index Python commencent à 0
            current_border = cell.border
            cell.border = Border(
                left=current_border.left,
                right=Side(style='thin', color='000000'),
                top=current_border.top,
                bottom=current_border.bottom
            )

# 💾 Sauvegarde du fichier final
now = datetime.now()
timestamp = now.strftime("%Y-%m-%d_%H-%M")

# Dossier de destination
export_folder = os.path.abspath(os.path.join(script_folder, "..", "Releves", "Syntheses_par_categorie", "Ressources"))
os.makedirs(export_folder, exist_ok=True)

# Nom du fichier final
filename = f"Synthese_Prix_Ressources_MAJ_{timestamp}.xlsx"
export_path = os.path.join(export_folder, filename)

wb_synth.save(export_path)
print(f"✅ Fichier exporté : {export_path}")
print("\n✅ Fusion terminée avec succès.")
input("Appuyez sur une touche pour quitter...")
